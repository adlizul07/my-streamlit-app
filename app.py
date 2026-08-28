import streamlit as st
import pandas as pd
import numpy as np
import io
import requests
import json
import re

from io import BytesIO
from bs4 import BeautifulSoup
import pdfplumber
from deep_translator import GoogleTranslator
from sentence_transformers import SentenceTransformer
from sklearn.cluster import AgglomerativeClustering
from sklearn.preprocessing import normalize
from openpyxl import load_workbook


# ============================================================
# CONFIG
# ============================================================

st.set_page_config(
    page_title="Insights Copilot",
    layout="wide",
    page_icon="\U0001F4CA",
    initial_sidebar_state="expanded",
)

STEPS = [
    ("step1", "Extract text from links"),
    ("step2", "Combine columns"),
    ("step3", "Remove duplicates"),
    ("step4", "Translate to English"),
    ("step5", "Semantic clustering"),
    ("step6", "AI intelligence"),
]

DEFAULT_OLLAMA_URL = "http://localhost:11434"

AI_FIELDS = [
    "pillar",
    "sub_pillar",
    "topic",
    "topic_summary",
    "spokesperson",
    "organisation",
    "location",
    "sentiment",
    "key_message",
    "issue",
    "campaign_or_initiative",
    "product_or_service",
    "competitor",
    "media_angle",
    "confidence",
]

# field name -> output column name
AI_COLUMNS = {
    "pillar": "AI_Pillar",
    "sub_pillar": "AI_Sub_Pillar",
    "topic": "AI_Topic",
    "topic_summary": "AI_Topic_Summary",
    "spokesperson": "AI_Spokesperson",
    "organisation": "AI_Organisation",
    "location": "AI_Location",
    "sentiment": "AI_Sentiment",
    "key_message": "AI_Key_Message",
    "issue": "AI_Issue",
    "campaign_or_initiative": "AI_Campaign_or_Initiative",
    "product_or_service": "AI_Product_or_Service",
    "competitor": "AI_Competitor",
    "media_angle": "AI_Media_Angle",
    "confidence": "AI_Confidence",
}


# ============================================================
# STYLES
# ============================================================

st.markdown("""
<style>

@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;600;700&display=swap');

:root {
    --bg-base:        #0d0f12;
    --bg-raised:      #13161b;
    --border:         #252932;
    --border-strong:  #333846;
    --accent:         #2e7dff;
    --accent-soft:    #2e7dff18;
    --ok:             #00d48f;
    --warn:           #f5a623;
    --err:            #ff4757;
    --muted:          #4a5162;
    --text:           #e8eaf0;
    --text-2:         #8b92a5;
    --radius:         8px;
}

html, body, [data-testid="stAppViewContainer"] {
    background-color: var(--bg-base) !important;
    font-family: 'DM Sans', sans-serif;
    color: var(--text);
}

[data-testid="stSidebar"] {
    background-color: var(--bg-raised) !important;
    border-right: 1px solid var(--border);
}

#MainMenu, footer, header { visibility: hidden; }
[data-testid="stDecoration"] { display: none; }

.block-container {
    padding: 2rem 3rem 4rem !important;
    max-width: 1180px;
}

.app-header {
    padding: 8px 0 20px;
    border-bottom: 1px solid var(--border);
    margin-bottom: 24px;
}

.app-header-title {
    font-size: 22px;
    font-weight: 700;
    letter-spacing: -0.4px;
}

.app-header-sub {
    font-size: 13px;
    color: var(--text-2);
    margin-top: 2px;
}

.pipeline-progress { display: flex; gap: 4px; margin: 4px 0 6px; }
.pipeline-step { flex: 1; height: 4px; border-radius: 2px; background: var(--border); }
.pipeline-step.done { background: var(--ok); }
.pipeline-step.active { background: var(--accent); }
.pipeline-step.error { background: var(--err); }
.pipeline-caption { font-size: 12px; color: var(--muted); margin-bottom: 22px; }

.status-pill {
    display: inline-flex;
    align-items: center;
    gap: 6px;
    font-size: 12px;
    padding: 3px 10px;
    border-radius: 100px;
    border: 1px solid;
    margin-top: 10px;
}

.status-not-run { color: var(--warn); border-color: #f5a62355; background: #f5a62310; }
.status-running { color: var(--accent); border-color: #2e7dff55; background: var(--accent-soft); }
.status-done { color: var(--ok); border-color: #00d48f55; background: #00d48f10; }
.status-error { color: var(--err); border-color: #ff475755; background: #ff475710; }

.section-label {
    font-size: 12px;
    font-weight: 600;
    color: var(--text-2);
    margin-bottom: 10px;
}

.step-note { color: var(--text-2); font-size: 13px; margin-bottom: 14px; }

[data-testid="stExpander"] {
    border: 1px solid var(--border) !important;
    border-radius: var(--radius) !important;
    background: var(--bg-base) !important;
    margin-bottom: 8px;
}

[data-testid="stExpander"] summary {
    font-weight: 500 !important;
    font-size: 14px !important;
}

.stButton > button {
    font-family: 'DM Sans', sans-serif !important;
    font-weight: 500 !important;
    font-size: 13px !important;
    border-radius: var(--radius) !important;
    padding: 7px 18px !important;
    border: 1px solid var(--border-strong) !important;
    background: var(--bg-base) !important;
    color: var(--text) !important;
    box-shadow: none !important;
}

.stButton > button:hover {
    border-color: var(--accent) !important;
    color: var(--accent) !important;
}

.stButton > button[kind="primary"] {
    background: var(--accent) !important;
    border-color: var(--accent) !important;
    color: #fff !important;
}

[data-testid="stDownloadButton"] > button {
    width: 100%;
    background: var(--accent) !important;
    border: 1px solid var(--accent) !important;
    color: #fff !important;
    font-weight: 600 !important;
    padding: 10px 20px !important;
    border-radius: var(--radius) !important;
}

.stTextInput > div > div > input,
.stTextArea > div > div > textarea,
.stSelectbox > div > div,
.stMultiSelect > div > div {
    background: var(--bg-base) !important;
    border: 1px solid var(--border-strong) !important;
    border-radius: var(--radius) !important;
    color: var(--text) !important;
}

.stTextInput > div > div > input:focus { border-color: var(--accent) !important; }

.stSelectbox div[data-baseweb="select"] *,
.stMultiSelect div[data-baseweb="select"] * { color: var(--text) !important; }

.stSelectbox div[data-baseweb="select"] svg,
.stMultiSelect div[data-baseweb="select"] svg { fill: var(--text-2) !important; }

.stMultiSelect [data-baseweb="tag"] {
    background: var(--accent-soft) !important;
    border: 1px solid #2e7dff55 !important;
}

.stMultiSelect [data-baseweb="tag"] span { color: var(--accent) !important; }

div[data-baseweb="popover"] ul,
div[data-baseweb="popover"] li {
    background: var(--bg-base) !important;
    color: var(--text) !important;
}

div[data-baseweb="popover"] li:hover { background: var(--accent-soft) !important; }

label,
.stSelectbox label,
.stMultiSelect label,
.stTextInput label,
.stCheckbox label,
.stSlider label {
    color: var(--text-2) !important;
    font-size: 13px !important;
    font-weight: 500 !important;
}

[data-testid="stDataFrame"] {
    border: 1px solid var(--border) !important;
    border-radius: var(--radius) !important;
    overflow: hidden !important;
}

[data-testid="stMetricValue"] { font-size: 20px !important; font-weight: 600 !important; }
[data-testid="stMetricLabel"] { color: var(--text-2) !important; }

[data-testid="stProgress"] > div > div > div > div { background-color: var(--accent) !important; }

hr { border-color: var(--border) !important; }
.stAlert { border-radius: var(--radius) !important; }

</style>
""", unsafe_allow_html=True)


# ============================================================
# SESSION STATE
# ============================================================

if "data" not in st.session_state:
    st.session_state.data = None

if "status" not in st.session_state:
    st.session_state.status = {k: "Not run" for k, _ in STEPS}

if "undo" not in st.session_state:
    st.session_state.undo = None

if "export_buffer" not in st.session_state:
    st.session_state.export_buffer = None


# ============================================================
# STATUS HELPERS
# ============================================================

def set_status(step, value):
    st.session_state.status[step] = value


def get_status(step):
    return st.session_state.status.get(step, "Not run")


def status_pill(step):
    s = get_status(step)

    cls = {
        "Not run": "status-not-run",
        "Running": "status-running",
        "Done": "status-done",
        "Error": "status-error",
    }.get(s, "status-not-run")

    dot = {
        "Not run": "\u25CB",
        "Running": "\u25C9",
        "Done": "\u25CF",
        "Error": "\u2715",
    }.get(s, "\u25CB")

    return f'<span class="status-pill {cls}">{dot} {s}</span>'


def save_undo(label):
    st.session_state.undo = (
        label,
        st.session_state.data.copy(deep=True),
        dict(st.session_state.status),
    )


# ============================================================
# SENTENCE TRANSFORMER
# ============================================================

@st.cache_resource
def load_model():
    return SentenceTransformer("paraphrase-multilingual-mpnet-base-v2")


# ============================================================
# GENERAL HELPERS
# ============================================================

_UA = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}


def clean_text(text):
    if pd.isnull(text):
        return text
    return str(text).replace("_x000D_", " ").replace("\n", " ").strip()


def _html_text(html):
    soup = BeautifulSoup(html, "html.parser")

    for tag in soup(["script", "style", "noscript", "header", "footer", "nav"]):
        tag.decompose()

    return " ".join(soup.get_text(separator=" ").split())


def _pdf_text(content):
    parts = []

    with pdfplumber.open(io.BytesIO(content)) as pdf:
        for page in pdf.pages:
            parts.append(page.extract_text() or "")

    return " ".join(" ".join(parts).split())


def extract_from_link(url):
    if not url or not isinstance(url, str):
        return "Link broken"

    try:
        r = requests.get(url, headers=_UA, timeout=25)

        if r.status_code != 200:
            return "Link broken"

        ctype = r.headers.get("Content-Type", "").lower()

        if "pdf" in ctype or url.lower().endswith(".pdf"):
            text = _pdf_text(r.content)
        elif "html" in ctype or "text" in ctype:
            text = _html_text(r.text)
        else:
            return "Link broken"

        text = text.strip()
        return text if text else "Link broken"

    except Exception:
        return "Link broken"


# ============================================================
# EXCEL
# ============================================================

def load_excel(file, sheet):
    wb = load_workbook(file, data_only=False)
    ws = wb[sheet]

    headers = [cell.value for cell in ws[1]]
    cleaned_headers = []

    for h in headers:
        if h is None or str(h).strip() == "":
            st.error("A column header is blank. Fix the header row and upload again.")
            st.stop()
        cleaned_headers.append(str(h).strip())

    if len(cleaned_headers) != len(set(cleaned_headers)):
        st.error("Two columns share the same header. Rename them and upload again.")
        st.stop()

    rows = [[cell.value for cell in row] for row in ws.iter_rows(min_row=2)]
    df = pd.DataFrame(rows, columns=cleaned_headers)

    # Capture Headline hyperlinks
    if "Headline" in df.columns:
        df["Headline_Link"] = None
        headline_col_idx = list(df.columns).index("Headline")

        for i, row in enumerate(ws.iter_rows(min_row=2)):
            cell = row[headline_col_idx]
            if cell.hyperlink:
                df.loc[i, "Headline_Link"] = cell.hyperlink.target

    return df


def to_excel(df):
    buffer = BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Media Intelligence")
        sheet = writer.sheets["Media Intelligence"]

        # Restore Headline hyperlinks
        if "Headline" in df.columns and "Headline_Link" in df.columns:
            col_idx = list(df.columns).index("Headline") + 1

            for row_idx in range(len(df)):
                link = df.iloc[row_idx]["Headline_Link"]

                if (
                    pd.notna(link)
                    and isinstance(link, str)
                    and link.startswith(("http://", "https://"))
                ):
                    cell = sheet.cell(row=row_idx + 2, column=col_idx)
                    cell.hyperlink = link
                    cell.style = "Hyperlink"

        # Remove internal helper column
        if "Headline_Link" in df.columns:
            helper_col = list(df.columns).index("Headline_Link") + 1
            sheet.delete_cols(helper_col)

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

        # Column widths, sampled from the first 100 rows
        for column_cells in sheet.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter

            for cell in column_cells[:100]:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass

            sheet.column_dimensions[column_letter].width = min(
                max(max_length + 2, 10), 45
            )

    buffer.seek(0)
    return buffer


# ============================================================
# CHUNKING
# ============================================================

def chunk_bounds(total, chunks=25):
    if total == 0:
        return []

    size = max(1, -(-total // chunks))
    return [(i, min(i + size, total)) for i in range(0, total, size)]


# ============================================================
# STEP 1 — EXTRACT
# ============================================================

def run_extract(df, media_col, extract_col, allowed_types):
    if "Headline_Link" not in df.columns:
        raise ValueError("This file has no hyperlinks on the Headline column.")

    allowed_lower = {a.lower() for a in allowed_types}

    targets = [
        i for i in df.index
        if str(df.at[i, media_col]).strip().lower() in allowed_lower
        and (
            pd.isnull(df.at[i, extract_col])
            or str(df.at[i, extract_col]).strip() == ""
        )
    ]

    bar = st.progress(0, text="Extracting \u2014 0%")
    broken = 0

    for n, i in enumerate(targets):
        result = extract_from_link(df.at[i, "Headline_Link"])
        df.at[i, extract_col] = result

        if result == "Link broken":
            broken += 1

        pct = int(((n + 1) / max(1, len(targets))) * 100)
        bar.progress(
            pct,
            text=f"Extracting \u2014 {pct}% ({n + 1} of {len(targets)})",
        )

    bar.empty()

    return df, (
        f"Processed {len(targets)} rows \u2014 "
        f"{len(targets) - broken} extracted, {broken} link broken"
    )


# ============================================================
# STEP 2 — COMBINE
# ============================================================

def run_combine(df, cols):
    if not cols:
        raise ValueError("Pick at least one column to combine.")

    total = len(df)
    bar = st.progress(0, text="Combining \u2014 0%")

    pieces = []
    bounds = chunk_bounds(total)

    for n, (start, end) in enumerate(bounds):
        block = (
            df.iloc[start:end][cols]
            .fillna("")
            .astype(str)
            .agg(" ".join, axis=1)
        )
        pieces.append(block.map(clean_text))

        pct = int(((n + 1) / len(bounds)) * 100)
        bar.progress(pct, text=f"Combining \u2014 {pct}%")

    df["Combined"] = pd.concat(pieces) if pieces else ""
    bar.empty()

    return df, f"Combined {len(cols)} column(s) across {total} rows"


# ============================================================
# STEP 3 — DEDUPE
# ============================================================

def run_dedupe(df, exclude_cols):
    before = len(df)
    bar = st.progress(0, text="Comparing rows \u2014 0%")

    check_cols = [c for c in df.columns if c not in exclude_cols]

    bar.progress(50, text="Comparing rows \u2014 50%")
    df = df.drop_duplicates(subset=check_cols)
    bar.progress(100, text="Comparing rows \u2014 100%")
    bar.empty()

    removed = before - len(df)
    return df, f"Removed {removed} duplicate row(s) \u2014 {len(df)} remaining"


# ============================================================
# STEP 4 — TRANSLATE
# ============================================================

def run_translate(df):
    if "Combined" not in df.columns:
        raise ValueError("Run 'Combine columns' first.")

    translator = GoogleTranslator(source="auto", target="en")
    texts = df["Combined"].astype(str).tolist()
    total = len(texts)

    bar = st.progress(0, text="Translating \u2014 0%")
    out = []
    failed = 0

    for n, t in enumerate(texts):
        try:
            out.append(translator.translate(t[:2000]))
        except Exception:
            out.append(t)
            failed += 1

        pct = int(((n + 1) / max(1, total)) * 100)
        bar.progress(
            pct,
            text=f"Translating \u2014 {pct}% ({n + 1} of {total})",
        )

    df["Translated"] = out
    bar.empty()

    note = f"Translated {total - failed} of {total} rows"
    if failed:
        note += f" \u2014 {failed} kept in the original language"

    return df, note


# ============================================================
# STEP 5 — CLUSTERING
# ============================================================

def run_cluster(df, threshold):
    if "Combined" not in df.columns:
        raise ValueError("Run 'Combine columns' first.")

    model = load_model()
    texts = df["Combined"].astype(str).tolist()
    total = len(texts)

    if total < 2:
        raise ValueError("At least two rows are required for clustering.")

    batch_size = 64
    bar = st.progress(0, text="Generating embeddings \u2014 0%")
    embeddings = []

    for i in range(0, total, batch_size):
        batch = texts[i:i + batch_size]

        embeddings.extend(
            model.encode(batch, convert_to_numpy=True, show_progress_bar=False)
        )

        pct = min(100, int(((i + len(batch)) / total) * 100))
        bar.progress(pct, text=f"Generating embeddings \u2014 {pct}%")

    bar.progress(100, text="Grouping articles")

    emb = normalize(np.array(embeddings))

    clustering = AgglomerativeClustering(
        n_clusters=None,
        metric="cosine",
        linkage="average",
        distance_threshold=threshold,
    )

    df["Cluster"] = clustering.fit_predict(emb)

    summary = {}
    for cluster in sorted(df["Cluster"].unique()):
        cluster_rows = df[df["Cluster"] == cluster]
        summary[cluster] = " | ".join(
            cluster_rows["Combined"].head(3).astype(str).tolist()
        )

    df["Cluster_Description"] = df["Cluster"].map(summary)
    bar.empty()

    return df, f"Found {df['Cluster'].nunique()} clusters across {total} rows"


# ============================================================
# OLLAMA
# ============================================================

def check_ollama(base_url, model):
    try:
        response = requests.get(f"{base_url}/api/tags", timeout=5)

        if response.status_code != 200:
            return False, f"Ollama responded with HTTP {response.status_code}"

        models = response.json().get("models", [])
        model_names = [m.get("name", "") for m in models]

        found = any(
            name == model or name.startswith(model + ":")
            for name in model_names
        )

        if not found:
            available = ", ".join(model_names) if model_names else "none"
            return False, (
                f"Model '{model}' was not found in Ollama. "
                f"Installed models: {available}"
            )

        return True, "Ollama ready"

    except Exception as e:
        return False, (
            "Cannot connect to Ollama. Make sure Ollama is running. "
            f"Details: {e}"
        )


def clean_json_response(text):
    """Ollama should return JSON, but models sometimes wrap it in fences."""
    if not text:
        return None

    text = text.strip()
    text = re.sub(r"^```json\s*", "", text, flags=re.IGNORECASE)
    text = re.sub(r"^```\s*", "", text)
    text = re.sub(r"\s*```$", "", text)
    text = text.strip()

    try:
        return json.loads(text)
    except Exception:
        pass

    start = text.find("{")
    end = text.rfind("}")

    if start >= 0 and end > start:
        try:
            return json.loads(text[start:end + 1])
        except Exception:
            pass

    return None


def ollama_generate(base_url, model, prompt, timeout=600):
    response = requests.post(
        f"{base_url}/api/generate",
        json={
            "model": model,
            "prompt": prompt,
            "stream": False,
            "format": "json",
            "options": {"temperature": 0.1},
        },
        timeout=timeout,
    )

    response.raise_for_status()

    raw = response.json().get("response", "")
    parsed = clean_json_response(raw)

    if parsed is None:
        raise ValueError("Ollama returned invalid JSON:\n" + raw[:1000])

    return parsed


# ============================================================
# AI PROMPT
# ============================================================

def build_cluster_prompt(cluster_id, articles):
    article_text = []

    for n, article in enumerate(articles, start=1):
        headline = str(article.get("Headline", ""))

        translated = str(
            article.get("Translated", article.get("Combined", ""))
        )[:3000]

        article_text.append(
            f"\nARTICLE {n}\n\nHEADLINE:\n{headline}\n\nCONTENT:\n{translated}\n"
        )

    articles_block = "\n".join(article_text)

    prompt = f"""
You are a senior media intelligence analyst.

Analyse the following media articles from CLUSTER {cluster_id}.

The articles have already been grouped together using semantic
similarity. Your job is to identify the common media narrative
and intelligence represented by this cluster.

IMPORTANT RULES:

1. Do NOT invent facts.
2. Do NOT invent spokesperson names.
3. Only identify a spokesperson if a person's name is explicitly
   present in the supplied articles.
4. If no spokesperson is clearly identified, return
   "Not identified".
5. Identify the most appropriate broad PILLAR from the content.
6. The pillar is NOT provided by a codebook. Infer it yourself.
7. Keep the pillar broad and reusable across many clusters.
8. The topic should be specific to this cluster.
9. Topic names must be straightforward and descriptive rather
   than headline-style or punchy.
10. Sentiment should describe the overall media tone.
11. Use Positive, Neutral, Negative or Mixed.
12. If several organisations are mentioned, identify the main
    organisation being discussed.
13. Locations should only be included if explicitly mentioned.
14. Key message must be a concise factual statement.
15. Confidence should reflect how clearly the articles support
    the classification.

Return ONLY valid JSON.

Required JSON structure:

{{
    "pillar": "",
    "sub_pillar": "",
    "topic": "",
    "topic_summary": "",
    "spokesperson": "",
    "organisation": "",
    "location": "",
    "sentiment": "",
    "key_message": "",
    "issue": "",
    "campaign_or_initiative": "",
    "product_or_service": "",
    "competitor": "",
    "media_angle": "",
    "confidence": 0.0
}}

CLUSTER ARTICLES:

{articles_block}
"""

    return prompt


# ============================================================
# STEP 6 — AI INTELLIGENCE
# ============================================================

def run_ai_intelligence(df, base_url, model_name, articles_per_cluster):
    if "Cluster" not in df.columns:
        raise ValueError("Run Semantic clustering first.")

    if "Translated" not in df.columns and "Combined" not in df.columns:
        raise ValueError("Run Combine columns first.")

    ok, message = check_ollama(base_url, model_name)
    if not ok:
        raise ValueError(message)

    clusters = sorted(df["Cluster"].dropna().unique())
    total_clusters = len(clusters)

    bar = st.progress(0, text="Preparing AI analysis...")

    results = {}
    failed_clusters = []

    text_col = "Translated" if "Translated" in df.columns else "Combined"

    for n, cluster_id in enumerate(clusters, start=1):
        cluster_df = df[df["Cluster"] == cluster_id].copy()

        # Send the longest articles, which usually carry the most context
        cluster_df["_text_length"] = (
            cluster_df[text_col].fillna("").astype(str).str.len()
        )

        representatives = cluster_df.sort_values(
            "_text_length", ascending=False
        ).head(articles_per_cluster)

        articles = representatives.drop(
            columns=["_text_length"], errors="ignore"
        ).to_dict(orient="records")

        prompt = build_cluster_prompt(cluster_id, articles)

        try:
            result = ollama_generate(base_url, model_name, prompt)

            cleaned = {}
            for field in AI_FIELDS:
                value = result.get(field, "")
                cleaned[field] = "" if value is None else value

            try:
                cleaned["confidence"] = float(cleaned["confidence"])
            except Exception:
                cleaned["confidence"] = 0.0

            cleaned["confidence"] = max(0.0, min(1.0, cleaned["confidence"]))
            results[cluster_id] = cleaned

        except Exception as e:
            failed_clusters.append((cluster_id, str(e)))
            results[cluster_id] = {
                field: (0.0 if field == "confidence" else "Not identified")
                for field in AI_FIELDS
            }

        pct = int((n / max(1, total_clusters)) * 100)
        bar.progress(
            pct,
            text=f"Ollama analysing clusters \u2014 {pct}% ({n} of {total_clusters})",
        )

    # Map cluster intelligence back to every article
    for column in AI_COLUMNS.values():
        df[column] = ""

    for cluster_id, result in results.items():
        mask = df["Cluster"] == cluster_id

        for field, column in AI_COLUMNS.items():
            df.loc[mask, column] = result[field]

    bar.empty()

    failed_count = len(failed_clusters)

    if failed_count:
        note = (
            f"Analysed {total_clusters - failed_count} of "
            f"{total_clusters} clusters. {failed_count} clusters failed."
        )
    else:
        note = (
            f"Ollama analysed all {total_clusters} clusters and mapped "
            f"the intelligence back to {len(df):,} articles."
        )

    return df, note


# ============================================================
# STEP MAP
# ============================================================

STEP_FUNCTIONS = {
    "step1": run_extract,
    "step2": run_combine,
    "step3": run_dedupe,
    "step4": run_translate,
    "step5": run_cluster,
    "step6": run_ai_intelligence,
}


# ============================================================
# EXECUTE
# ============================================================

def execute(step_key, fn, *args):
    label = dict(STEPS)[step_key]

    save_undo(label)
    set_status(step_key, "Running")

    try:
        new_df, note = fn(st.session_state.data, *args)

        st.session_state.data = new_df
        st.session_state.export_buffer = None

        set_status(step_key, "Done")
        st.success(note)
        return True

    except Exception as e:
        set_status(step_key, "Error")
        st.error(str(e))
        return False


# ============================================================
# HEADER
# ============================================================

st.markdown("""
<div class="app-header">
<div class="app-header-title">Insights Copilot</div>
<div class="app-header-sub">Media intelligence pipeline \u00B7 cleaning, translation, clustering and AI intelligence</div>
</div>
""", unsafe_allow_html=True)


# ============================================================
# SIDEBAR
# ============================================================

with st.sidebar:

    st.markdown(
        '<div class="section-label">Data source</div>',
        unsafe_allow_html=True,
    )

    file = st.file_uploader("Excel file (.xlsx)", type=["xlsx"])

    if file is not None:
        sheet = st.selectbox("Sheet", pd.ExcelFile(file).sheet_names)

        if st.button("Load data", type="primary", use_container_width=True):
            st.session_state.data = load_excel(file, sheet)
            st.session_state.status = {k: "Not run" for k, _ in STEPS}
            st.session_state.undo = None
            st.session_state.export_buffer = None
            st.rerun()

    if st.session_state.data is not None:
        df_side = st.session_state.data

        st.markdown("---")
        st.markdown(
            '<div class="section-label">Step settings</div>',
            unsafe_allow_html=True,
        )

        with st.expander("Extract text from links", expanded=False):
            media_col = st.selectbox(
                "Media type column",
                df_side.columns,
                index=(
                    list(df_side.columns).index("Media Type")
                    if "Media Type" in df_side.columns
                    else 0
                ),
            )

            extract_col = st.selectbox(
                "Extract text column",
                df_side.columns,
                index=(
                    list(df_side.columns).index("Extract Text")
                    if "Extract Text" in df_side.columns
                    else 0
                ),
            )

            allowed_types = st.multiselect(
                "Media types to process",
                ["Online", "Newspaper", "TV", "Radio"],
                default=["Online", "Newspaper"],
            )

        with st.expander("Combine columns", expanded=False):
            combine_cols = st.multiselect("Columns to combine", df_side.columns)

        with st.expander("Remove duplicates", expanded=False):
            exclude_cols = st.multiselect(
                "Columns to ignore when comparing", df_side.columns
            )

        with st.expander("Semantic clustering", expanded=False):
            threshold = st.slider(
                "Distance threshold (lower is stricter)",
                0.25,
                0.35,
                0.28,
                step=0.01,
            )

        with st.expander("AI intelligence \u2014 Ollama", expanded=True):
            ollama_url = st.text_input("Ollama URL", DEFAULT_OLLAMA_URL)

            ollama_model = st.text_input(
                "Ollama model",
                "llama3.1:8b",
                help="Use a model already installed in Ollama.",
            )

            articles_per_cluster = st.slider(
                "Articles sampled per cluster",
                min_value=2,
                max_value=10,
                value=5,
                help=(
                    "Ollama analyses representative articles from each "
                    "cluster. More articles improve context but take longer."
                ),
            )

            if st.button("Test Ollama connection", use_container_width=True):
                ok, message = check_ollama(ollama_url, ollama_model)

                if ok:
                    st.success(message)
                else:
                    st.error(message)

            st.caption(
                "No pillar or spokesperson codebook is required. "
                "Ollama infers the intelligence from the articles."
            )

        st.markdown("---")
        st.markdown(
            '<div class="section-label">Run everything</div>',
            unsafe_allow_html=True,
        )

        run_all_steps = st.multiselect(
            "Steps to include",
            [name for _, name in STEPS],
            default=[name for _, name in STEPS],
        )

        run_all = st.button(
            "Run selected steps", type="primary", use_container_width=True
        )

        st.markdown("---")

        undo_label = (
            st.session_state.undo[0] if st.session_state.undo else None
        )

        if st.button(
            f"Undo: {undo_label}" if undo_label else "Undo last step",
            disabled=undo_label is None,
            use_container_width=True,
        ):
            _, prev_df, prev_status = st.session_state.undo

            st.session_state.data = prev_df
            st.session_state.status = prev_status
            st.session_state.undo = None
            st.session_state.export_buffer = None
            st.rerun()

        if st.button("Start over", use_container_width=True):
            st.session_state.data = None
            st.session_state.status = {k: "Not run" for k, _ in STEPS}
            st.session_state.undo = None
            st.session_state.export_buffer = None
            st.rerun()

    else:
        run_all = False
        run_all_steps = []


# ============================================================
# DATA CHECK
# ============================================================

df = st.session_state.data

if df is None:
    st.info("Upload an Excel file in the sidebar to start.")
    st.stop()


# ============================================================
# PIPELINE STRIP
# ============================================================

strip = "".join(
    '<div class="pipeline-step %s"></div>'
    % {"Done": "done", "Running": "active", "Error": "error"}.get(
        get_status(k), ""
    )
    for k, _ in STEPS
)

done_count = sum(1 for k, _ in STEPS if get_status(k) == "Done")

st.markdown(
    f'<div class="pipeline-progress">{strip}</div>'
    f'<div class="pipeline-caption">{done_count} of {len(STEPS)} steps done</div>',
    unsafe_allow_html=True,
)


# ============================================================
# METRICS
# ============================================================

m1, m2, m3 = st.columns(3)

m1.metric("Rows", f"{len(df):,}")

m2.metric(
    "Columns",
    len([c for c in df.columns if c != "Headline_Link"]),
)

m3.metric("Steps done", f"{done_count} / {len(STEPS)}")


# ============================================================
# PREVIEW
# ============================================================

st.markdown(
    '<div class="section-label" style="margin-top:18px;">Preview</div>',
    unsafe_allow_html=True,
)

st.dataframe(df.head(10), use_container_width=True)

st.markdown("<br>", unsafe_allow_html=True)

st.markdown(
    '<div class="section-label">Steps</div>',
    unsafe_allow_html=True,
)


# ============================================================
# RUN ALL
# ============================================================

if run_all:
    name_to_key = {name: key for key, name in STEPS}
    selected_names = [n for _, n in STEPS if n in run_all_steps]

    for name in selected_names:
        key = name_to_key[name]
        st.markdown(f"**{name}**")

        if key == "step1":
            ok = execute(key, run_extract, media_col, extract_col, allowed_types)
        elif key == "step2":
            ok = execute(key, run_combine, combine_cols)
        elif key == "step3":
            ok = execute(key, run_dedupe, exclude_cols)
        elif key == "step4":
            ok = execute(key, run_translate)
        elif key == "step5":
            ok = execute(key, run_cluster, threshold)
        else:
            ok = execute(
                key,
                run_ai_intelligence,
                ollama_url,
                ollama_model,
                articles_per_cluster,
            )

        if not ok:
            st.warning("Stopped here. Fix the setting above and run again.")
            break

    df = st.session_state.data


# ============================================================
# STEP HEADER
# ============================================================

def step_header(key):
    name = dict(STEPS)[key]
    mark = {"Done": "\u2713 ", "Error": "\u2715 "}.get(get_status(key), "")
    return f"{mark}{name}"


# ============================================================
# STEP 1 UI
# ============================================================

with st.expander(
    step_header("step1"),
    expanded=get_status("step1") != "Done",
):
    st.markdown(
        '<p class="step-note">Fills blank Extract Text cells by pulling '
        'the text behind the Headline link.</p>',
        unsafe_allow_html=True,
    )

    if st.button("Run", key="run1"):
        execute("step1", run_extract, media_col, extract_col, allowed_types)

    st.markdown(status_pill("step1"), unsafe_allow_html=True)


# ============================================================
# STEP 2 UI
# ============================================================

with st.expander(
    step_header("step2"),
    expanded=get_status("step2") != "Done",
):
    st.markdown(
        '<p class="step-note">Joins the selected columns into one '
        'Combined field.</p>',
        unsafe_allow_html=True,
    )

    if st.button("Run", key="run2"):
        execute("step2", run_combine, combine_cols)

    st.markdown(status_pill("step2"), unsafe_allow_html=True)


# ============================================================
# STEP 3 UI
# ============================================================

with st.expander(
    step_header("step3"),
    expanded=get_status("step3") != "Done",
):
    st.markdown(
        '<p class="step-note">Drops rows that match on every column '
        'apart from the ones you excluded.</p>',
        unsafe_allow_html=True,
    )

    if st.button("Run", key="run3"):
        execute("step3", run_dedupe, exclude_cols)

    st.markdown(status_pill("step3"), unsafe_allow_html=True)


# ============================================================
# STEP 4 UI
# ============================================================

with st.expander(
    step_header("step4"),
    expanded=get_status("step4") != "Done",
):
    st.markdown(
        '<p class="step-note">Translates the Combined field to English.</p>',
        unsafe_allow_html=True,
    )

    if st.button("Run", key="run4"):
        execute("step4", run_translate)

    st.markdown(status_pill("step4"), unsafe_allow_html=True)


# ============================================================
# STEP 5 UI
# ============================================================

with st.expander(
    step_header("step5"),
    expanded=get_status("step5") != "Done",
):
    st.markdown(
        '<p class="step-note">Groups similar articles using multilingual '
        'sentence embeddings and agglomerative clustering.</p>',
        unsafe_allow_html=True,
    )

    if st.button("Run", key="run5"):
        execute("step5", run_cluster, threshold)

    st.markdown(status_pill("step5"), unsafe_allow_html=True)


# ============================================================
# STEP 6 UI — OLLAMA
# ============================================================

with st.expander(
    step_header("step6"),
    expanded=get_status("step6") != "Done",
):
    st.markdown(
        '<p class="step-note">Uses Ollama to analyse each semantic cluster '
        'and identify pillars, sub-pillars, topics, spokespersons, sentiment, '
        'key messages, issues, organisations, locations and media angles. '
        'The intelligence is then mapped back to every article in the '
        'cluster.</p>',
        unsafe_allow_html=True,
    )

    if "Cluster" not in df.columns:
        st.warning("Run Semantic clustering first.")

    else:
        cluster_count = df["Cluster"].nunique()

        st.info(
            f"{len(df):,} articles have been grouped into "
            f"{cluster_count:,} clusters. Ollama will analyse the clusters "
            f"rather than every article individually."
        )

        if st.button("Run AI Intelligence", key="run6", type="primary"):
            execute(
                "step6",
                run_ai_intelligence,
                ollama_url,
                ollama_model,
                articles_per_cluster,
            )

    st.markdown(status_pill("step6"), unsafe_allow_html=True)


# ============================================================
# AI RESULTS PREVIEW
# ============================================================

if "AI_Pillar" in df.columns:

    st.markdown("---")
    st.markdown(
        '<div class="section-label">AI Intelligence Preview</div>',
        unsafe_allow_html=True,
    )

    ai_preview_cols = [
        "Cluster",
        "AI_Pillar",
        "AI_Sub_Pillar",
        "AI_Topic",
        "AI_Spokesperson",
        "AI_Sentiment",
        "AI_Key_Message",
        "AI_Confidence",
    ]

    available_ai_cols = [c for c in ai_preview_cols if c in df.columns]

    st.dataframe(df[available_ai_cols].head(20), use_container_width=True)

    a1, a2, a3, a4 = st.columns(4)

    a1.metric("Pillars", df["AI_Pillar"].replace("", np.nan).nunique())

    a2.metric("Topics", df["AI_Topic"].replace("", np.nan).nunique())

    a3.metric(
        "Spokespersons",
        df["AI_Spokesperson"]
        .replace("", np.nan)
        .replace("Not identified", np.nan)
        .nunique(),
    )

    a4.metric("AI analysed rows", f"{len(df):,}")


# ============================================================
# EXPORT
# ============================================================

st.markdown("---")

st.markdown(
    '<div class="section-label">Export</div>',
    unsafe_allow_html=True,
)

e1, e2 = st.columns([2, 1])

with e1:
    filename = st.text_input(
        "File name",
        "media_intelligence_output.xlsx",
        label_visibility="collapsed",
    )

with e2:
    if st.button("Prepare file", use_container_width=True):
        st.session_state.export_buffer = to_excel(
            st.session_state.data
        ).getvalue()

if st.session_state.export_buffer is not None:
    st.download_button(
        "Download Excel",
        data=st.session_state.export_buffer,
        file_name=filename,
        mime=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )

else:
    st.caption(
        "Prepare the file first, then the download button appears here."
    )